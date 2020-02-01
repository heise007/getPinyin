#! /usr/bin/env python
#coding:utf-8
#encoding=utf-8
import os
import time
import sys
# reload(sys)  
# sys.setdefaultencoding('utf8')

import xlrd
# import xlwt
from xpinyin import Pinyin
from openpyxl import load_workbook

if __name__ == '__main__':

    department_No = 1
    name_No = 2
    englishname_No = 2
    phonenumber_No = 4
    fromfile = './2.xlsx'
    tofile = './export.xlsx'

    wb = load_workbook(tofile)
    sheet_ranges = wb['personInfo']
    ws=wb['personInfo']
    
    
    # global logger
    pingyin=[]
    data= xlrd.open_workbook(fromfile) 
    table = data.sheets()[0]
    nrows = table.nrows
    print(nrows)
    p=Pinyin()
    
    wb.guess_types = True   #猜测格式类型
    ws=wb.active
    #成都大学>>
    for index in range(4, 100):
        ws['b'+str(index)]=''
        ws['d'+str(index)]=''
        ws['e'+str(index)]=''
        ws['g'+str(index)]=''
    wb.save(tofile)
    for index in range(1,nrows):
        #print index
        #print nrows
        #print table.nrows
        #print table.row(index)
        
        # 部门------------------------------------------------------------------------------------
        ws['b'+str(index+3)]='成都大学>>'+str(table.row(index)[department_No].value)
        print('成都大学>>'+str(table.row(index)[department_No].value))

        # 中文名------------------------------------------------------------------------------------
        ws['D'+str(index+3)]=str(table.row(index)[name_No].value).replace(" ","")
        print(str(table.row(index)[name_No].value).replace(" ",""))

        # 英文名------------------------------------------------------------------------------------
        if englishname_No == name_No:# 中文名字的列号 == 英文名的列号，就表示要从汉字转换得到拼音
            ws['e'+str(index+3)]=str(p.get_pinyin((str(table.row(index)[englishname_No].value))).replace(" ","").replace("-",""))
            print(p.get_pinyin((str(table.row(index)[englishname_No].value)),tone_marks='marks').replace(" ","").replace("-",""))
        else:# 中文名字的列号 != 英文名的列号，就表示表里面已经有了拼音，直接读过来
            ws['e'+str(index+3)]=str(table.row(index)[englishname_No].value).replace(" ","").replace("-","")
            print(p.get_pinyin((str(table.row(index)[englishname_No].value))).replace(" ","").replace("-",""))

       # 电话------------------------------------------------------------------------------------
        if type(table.row(index)[phonenumber_No].value) is float:
            print(int(table.row(index)[phonenumber_No].value))
            ws['G'+str(index+3)]=str(int(table.row(index)[phonenumber_No].value)).replace(" ","")
        else:
            print(str(table.row(index)[phonenumber_No].value).replace(" ",""))
            ws['G'+str(index+3)]=str(table.row(index)[phonenumber_No].value).replace(" ","")

    wb.save(tofile)

    '''
    filename="./pingyin.txt"
    with open(filename, 'w') as file_object:
        for index1 in range(0,len(pingyin)):
            file_object.write(pingyin[index1]+"\r\n")
    workbook = xlwt.Workbook(encoding = 'ascii')
    worksheet = workbook.add_sheet('sheet1')
    worksheet.write(0, 0, label = 'test')
    workbook.save('Myxls.xls')
    '''
